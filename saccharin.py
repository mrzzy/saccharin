#
# Saccharin
# blood sugar data cleaning & export
#

from typing import cast, List
import numpy as np
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd

from datetime import datetime
from argparse import ArgumentParser
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter, range_boundaries


## Feature Extraction & Engineering
def to_float(series: pd.Series, default: float) -> pd.Series:
    """Convert the given series to float, using the given default if conversion fails."""
    return pd.to_numeric(series, "coerce").fillna(default).astype(float)  # type: ignore


def drop_empty(df: pd.DataFrame) -> pd.DataFrame:
    """Drop fully empty columns from the given DataFrame"""
    # trim empty whitespace in string columns
    df = df.applymap((lambda s: s.strip() if isinstance(s, str) else s))

    # remove empty & nan coluns
    df = df.replace("", np.nan).dropna(axis=1, how="all")  # type: ignore

    return df


def sort_tags(tags: pd.Series) -> pd.Series:
    """Sort the comma seperated tags in the given series"""
    return (
        tags.str.split(",")
        .map(
            # list type check needed as sorted cannot sort discrete float values
            lambda tags: tags
            if not isinstance(tags, List)
            else sorted(tags)
        )
        .str.join(",")
    )


def read_sugar_df(csv_path: str) -> pd.DataFrame:
    """Read the blood sugar data from the given CSV as a DataFrame"""
    # read sugar data csv export
    col_dtypes = {
        "Date": str,
        "Time": str,
        "Tags": str,
        "Blood Sugar Measurement (mmol/L)": np.float64,
        "Basal Injection Units": np.float64,
        "Insulin (Meal)": np.float64,
        "Insulin (Correction)": np.float64,
        "Meal Carbohydrates (Grams, Factor 1)": np.float64,
        "Meal Descriptions": str,
        "Note": str,
    }

    sugar_df = pd.read_csv(
        csv_path,
        dtype=col_dtypes,
    )
    if not isinstance(sugar_df, pd.DataFrame):
        raise ValueError("Expected sugar_df to be a DataFrame")

    # parse date & time columns
    sugar_df["Date"] = pd.to_datetime(sugar_df["Date"]).apply((lambda dt: dt.date()))  # type: ignore
    sugar_df["Time"] = pd.to_datetime(sugar_df["Time"]).apply((lambda dt: dt.time()))  # type: ignore
    # ensure identical sets of tags are represented by the same string
    sugar_df["Tags"] = sort_tags(sugar_df["Tags"])
    return drop_empty(sugar_df)


def label_outlier(sugar_level: float, outlier_high: float, outlier_low: float) -> str:
    """Label if the given sugar level is an outlier: outside outlier high & low constraints."""
    if sugar_level > outlier_high:
        return "High"
    elif sugar_level < outlier_low:
        return "Low"
    else:
        return ""


## Excel Utilities
def fit_sheet_cols(worksheet: Worksheet):
    """Autofit the given worksheet's columns to content"""
    for col in worksheet.iter_cols():
        col_letter = col[0].column_letter
        col_width = max([len(str(cell.value)) for cell in col])
        worksheet.column_dimensions[col_letter].width = col_width * 1.05  # type: ignore
    return worksheet


def convert_table(worksheet: Worksheet) -> Table:
    """Convert the data in the given worksheet into a table"""
    name = worksheet.title.replace(" ", "_")
    data_range = worksheet.calculate_dimension()
    table = Table(displayName=name, ref=data_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    return table


def fill_conditional(
    worksheet: Worksheet, address: str, condition: str, color_hex: str
):
    """Apply conditional color fill the cell range specified by the given address in the given worksheet"""
    # remove hash sign in color_hex if present
    color = color_hex.replace("#", "")

    worksheet.conditional_formatting.add(  # type: ignore
        address,
        FormulaRule(
            formula=[condition],
            fill=PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            ),
        ),
    )


def template_excel(sugar_df: pd.DataFrame, stats_df: pd.DataFrame) -> Workbook:
    """Template the given dataframes into an Excel Workbook"""
    # template excel spreadsheet
    wb = Workbook()

    # copy sugar dataframe data into excel workbook
    sugar_ws = wb.active
    if sugar_ws is None:
        raise RuntimeError("Unexpected None from Workbook.active")
    date_prefix = f'{sugar_df["Date"].min().strftime("%m|%y")} - {sugar_df["Date"].max().strftime("%m|%y")}'
    sugar_ws.title = f"{date_prefix} Blood Sugar"

    bool_to_str = lambda has: "yes" if has else "no"
    sugar_df["Hyperglycemia"] = sugar_df["Hyperglycemia"].apply(bool_to_str)
    sugar_df["Hypoglycemia"] = sugar_df["Hypoglycemia"].apply(bool_to_str)
    for row in dataframe_to_rows(sugar_df, index=False, header=True):
        sugar_ws.append(row)
    fit_sheet_cols(sugar_ws)

    # copy summary statistics in excel workbook
    stats_ws = cast(Worksheet, wb.create_sheet(f"{date_prefix} Statistics"))
    stats_df = stats_df.applymap(
        (lambda stat: "" if np.isnan(stat) else "{:.1f}".format(stat))
    )
    for row in dataframe_to_rows(stats_df, index=True, header=True):
        stats_ws.append(row)
    stats_ws.delete_rows(2)
    fit_sheet_cols(stats_ws)

    # convert worksheets data into tables
    sugar_tbl = convert_table(sugar_ws)
    convert_table(stats_ws)

    # conditional formatting
    # calculate conditional formatting columns
    max_col = range_boundaries(sugar_tbl.ref)[2]
    hyper_col = get_column_letter(max_col - 2)
    hypo_col = get_column_letter(max_col - 1)
    outlier_col = get_column_letter(max_col)
    # highlight hyper and hypoglycemia
    fill_conditional(
        sugar_ws,
        address=str(sugar_tbl.ref),
        condition=f'OR(${hyper_col}1 = "yes", ${hypo_col}1 = "yes")',
        color_hex="FF7F7F",
    )
    # highlight outlier blood sugar levels
    fill_conditional(
        sugar_ws,
        address=str(sugar_tbl.ref),
        condition=f'OR(${outlier_col}1 = "High", ${outlier_col}1 = "Low")',
        color_hex="FFD580",
    )

    return wb


if __name__ == "__main__":
    # parse script command line args
    parser = ArgumentParser(description="Prepare blood sugar data for DDD evaluation")
    parser.add_argument("csv_path", help="Path to the blood sugar data CSV export.")
    parser.add_argument(
        "--start-from",
        help="Starting date in format DD/MM/YYYY. "
        "Data from the starting date onwards is included.",
        type=(lambda date_str: datetime.strptime(date_str, "%d/%m/%Y").date()),
    )
    parser.add_argument(
        "--out-xlsx", help="Path to write output Excel file", default="out.xlsx"
    )
    parser.add_argument(
        "--outlier-high",
        help="Upper blood sugar level limit in mmol/L to highlight as high level outlier.",
        default=9.5,
        type=float,
    )
    parser.add_argument(
        "--outlier-low",
        help="Lower blood sugar level limit in mmol/L to highlight as low outlier.",
        default=4.5,
        type=float,
    )
    args = parser.parse_args()

    # read blood sugar data
    sugar_df = read_sugar_df(args.csv_path)

    # filter data to only include entries from start date onwards
    if args.start_from is not None:
        sugar_df = sugar_df[sugar_df["Date"] >= args.start_from]

    # add hypo / hyperglycemia features
    sugar_df["Hyperglycemia"] = sugar_df["Blood Sugar Measurement (mmol/L)"] > 10.0
    sugar_df["Hypoglycemia"] = sugar_df["Blood Sugar Measurement (mmol/L)"] < 4.0

    # add insulin carb ratio (ICR) features
    sugar_df["Total Insulin (Meal)"] = to_float(
        sugar_df["Insulin (Meal)"], default=0.0
    ) + to_float(sugar_df["Insulin (Correction)"], default=0.0)

    sugar_df["Insulin Carb Ratio (ICR)"] = (
        sugar_df["Meal Carbohydrates (Grams, Factor 1)"]
        / sugar_df["Total Insulin (Meal)"]
    )

    # add outlier features
    sugar_df["Outlier"] = [
        label_outlier(
            sugar,
            args.outlier_high,
            args.outlier_low,
        )
        for sugar in sugar_df["Blood Sugar Measurement (mmol/L)"]
    ]

    # compute summary statistics
    stats_df = sugar_df.describe().drop(["25%", "75%"])
    if stats_df is None:
        raise RuntimeError("Unexpected None returned from DataFrame.drop()")
    stats_df = stats_df.rename(index={"50%": "median"})

    # compute hyperglycemia & hypoglycemia statistics
    for prefix in ["Hyper", "Hypo"]:
        glycemia = f"{prefix}glycemia"
        count = sugar_df[sugar_df[glycemia]][glycemia].agg("count")
        stats_df.loc["count", glycemia] = count

    # compute average blood sugar level by meal
    meal_tags = ["Breakfast", "Lunch", "Dinner", "Snack"]

    # filter blood sugar measurements with no tags (na = False)
    meal_sugar_df = sugar_df[
        sugar_df["Tags"].str.contains("|".join(meal_tags), na=False)
    ]
    meal_stats_df = (
        meal_sugar_df[["Blood Sugar Measurement (mmol/L)", "Tags"]]
        .groupby(by="Tags")
        .agg("mean")
        .rename(
            # add suffix to clarify that its blood sugar
            index=lambda s: f"{s} (mmol/L)",
            columns={"Blood Sugar Measurement (mmol/L)": "mean"},
        )
    )
    stats_df = stats_df.join(meal_stats_df.T)

    # template and save as excel file
    workbook = template_excel(sugar_df, stats_df)
    workbook.save(args.out_xlsx)
