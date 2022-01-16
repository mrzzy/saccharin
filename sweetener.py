#
# sweetener
# blood sugar data cleaning & export
#

from typing import cast
import numpy as np
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd

from datetime import datetime
from argparse import ArgumentParser
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows


def read_sugar_df(csv_path: str) -> pd.DataFrame:
    """Read the blood sugar data from the given CSV as a DataFrame"""
    # read sugar data csv export
    col_dtypes = {
        "Date": str,
        "Time": str,
        "Tags": str,
        "Blood Sugar Measurement (mmol/L)": np.float64,
        "Basal Injection Units": np.float64,
        "Insulin (Meal)": np.float64,
        "Insulin (Correction)": np.float64,
        "Meal Carbohydrates (Grams, Factor 1)": np.float64,
        "Meal Descriptions": str,
        "Note": str,
    }

    sugar_df = pd.read_csv(
        csv_path,
        usecols=list(col_dtypes.keys()),
        dtype=col_dtypes,
    )
    if not isinstance(sugar_df, pd.DataFrame):
        raise ValueError("Expected sugar_df to be a DataFrame")

    # parse date & time columns
    sugar_df["Date"] = pd.to_datetime(sugar_df["Date"]).apply((lambda dt: dt.date()))  # type: ignore
    sugar_df["Time"] = pd.to_datetime(sugar_df["Time"]).apply((lambda dt: dt.time()))  # type: ignore
    return sugar_df


def fit_sheet_cols(worksheet: Worksheet) -> Worksheet:
    """Autofit the given worksheet's columns to content"""
    for col in worksheet.iter_cols():
        col_letter = col[0].column_letter
        col_width = max([len(str(cell.value)) for cell in col])
        worksheet.column_dimensions[col_letter].width = col_width * 1.05
    return worksheet


def convert_table(worksheet: Worksheet) -> Worksheet:
    """Convert the data in the given worksheet into a table"""
    name = worksheet.title.replace(" ", "_")
    table = Table(displayName=name, ref=worksheet.calculate_dimension())
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    return worksheet


def template_excel(sugar_df: pd.DataFrame, stats_df: pd.DataFrame) -> Workbook:
    """Template the given dataframes into an Excel Workbook"""
    # template excel spreadsheet
    wb = Workbook()

    # copy sugar dataframe data into excel workbook
    bool_to_str = lambda has: "yes" if has else "no"
    sugar_df["Hyperglycemia"] = sugar_df["Hyperglycemia"].apply(bool_to_str)
    sugar_df["Hypoglycemia"] = sugar_df["Hypoglycemia"].apply(bool_to_str)
    sugar_ws = wb.active
    if sugar_ws is None:
        raise RuntimeError("Unexpected None from Workbook.active")

    date_prefix = f'{sugar_df["Date"].min().strftime("%m|%y")} - {sugar_df["Date"].max().strftime("%m|%y")}'
    sugar_ws.title = f"{date_prefix} Blood Sugar"
    for row in dataframe_to_rows(sugar_df, index=False, header=True):
        sugar_ws.append(row)
    fit_sheet_cols(sugar_ws)

    # copy summary statistics in excel workbook
    stats_ws = cast(Worksheet, wb.create_sheet(f"{date_prefix} Statistics"))
    stats_df = stats_df.applymap(
        (lambda stat: "" if np.isnan(stat) else "{:.1f}".format(stat))
    )
    for row in dataframe_to_rows(stats_df, index=True, header=True):
        stats_ws.append(row)
    stats_ws.delete_rows(2)
    fit_sheet_cols(stats_ws)

    # add tables on worksheets
    convert_table(sugar_ws)
    convert_table(stats_ws)

    return wb


if __name__ == "__main__":
    # parse script command line args
    parser = ArgumentParser(description="Prepare blood sugar data for DDD evaluation")
    parser.add_argument("csv_path", help="Path to the blood sugar data CSV export.")
    parser.add_argument(
        "--start-from",
        help="Starting date in format DD/MM/YYYY. "
        "Data from the starting date onwards is included.",
        type=(lambda date_str: datetime.strptime(date_str, "%d/%m/%Y").date()),
    )
    parser.add_argument(
        "--out-xlsx", help="Path to write output Excel file", default="out.xlsx"
    )
    args = parser.parse_args()

    # read blood sugar data
    sugar_df = read_sugar_df(args.csv_path)

    # filter data to only include entries from start date onwards
    if args.start_from is not None:
        sugar_df = sugar_df[sugar_df["Date"] >= args.start_from]

    # add hypo / hyperglycemia features
    sugar_df["Hyperglycemia"] = sugar_df["Blood Sugar Measurement (mmol/L)"] > 10.0
    sugar_df["Hypoglycemia"] = sugar_df["Blood Sugar Measurement (mmol/L)"] < 4.0

    # compute summary statistics
    stats_df = sugar_df.describe().drop(["25%", "75%"])
    if stats_df is None:
        raise RuntimeError("Unexpected None returned from DataFrame.drop()")
    stats_df = stats_df.rename(index={"50%": "median"})

    # compute hyperglycemia & hypoglycemia statistics
    for prefix in ["Hyper", "Hypo"]:
        glycemia = f"{prefix}glycemia"
        count = sugar_df[sugar_df[glycemia]][glycemia].agg("count")
        stats_df.loc["count", glycemia] = count

    # template and save as excel file
    workbook = template_excel(sugar_df, stats_df)
    workbook.save(args.out_xlsx)
